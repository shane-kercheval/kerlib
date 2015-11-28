import logging
import datetime
import os
import csv

import smtplib
import numpy
from PIL import ImageFont
from PIL import Image
from PIL import ImageDraw
from openpyxl import Workbook
from openpyxl.cell import get_column_letter


def send_email(username,
               password,
               subject,
               email_text,
               to_addresses,
               cc_addresses=list(),
               bcc_addresses=list(),
               test=False):
    # Sends a gmail email.
    if len(to_addresses) == 0:
        raise ValueError("'to_addresses' can't be empty.")
    if type(to_addresses) is not list:
        raise TypeError("to_addresses needs to be a list, but is a " +
                        type(to_addresses).__name__)
    if type(cc_addresses) is not list:
        raise TypeError("cc_addresses needs to be a list, but is a " +
                        type(cc_addresses).__name__)
    if type(bcc_addresses) is not list:
        raise TypeError("bcc_addresses needs to be a list, but is a " +
                        type(bcc_addresses).__name__)

    if not test:
        session = smtplib.SMTP('smtp.gmail.com', 587)
        session.ehlo()
        session.starttls()
        session.login(username, password)

    if len(cc_addresses) > 0:
        headers = "\n".join(["from: " + username,
                             "subject: " + subject,
                             "to: " + ", ".join(to_addresses),
                             "cc: " + ", ".join(cc_addresses),
                             "mime-version: 1.0",
                             "content-type: text/html"])
    else:
        headers = "\n".join(["from: " + username,
                             "subject: " + subject,
                             "to: " + ", ".join(to_addresses),
                             "mime-version: 1.0",
                             "content-type: text/html"])

    content = headers + "\n" + email_text
    send_to = list()
    send_to.extend(to_addresses)
    send_to.extend(cc_addresses)
    send_to.extend(bcc_addresses)

    logging.info("-----send_email (test={}; time='{}')"
                 .format(test, datetime.datetime.now()))
    logging.info("content: " + content)
    logging.info("sendto: " + ", ".join(send_to))
    logging.info("-----end send_email--------")

    if test:
        return content + "\n" + ", ".join(send_to)
    elif session is not None:
        session.sendmail(username, send_to, content)
        return True
    else:
        raise ValueError("'to_addresses' can't be empty.")


def create_image_with_text(image_path="hello.png", text="test"):
    image_width, image_height = (300, 200)
    with Image.new("RGBA", (image_width, image_height), "yellow") as image:
        image.save(image_path, "PNG")
    add_text_to_image(image_path, text, image_font_size=20, fontColor="red")


def add_text_to_image(image_path, text,
                      new_image_path="newImage.png",
                      image_type="PNG",
                      image_font_size=30,
                      font_color="white"):

    with Image.open(image_path) as image:
        draw = ImageDraw.Draw(image)
        image_width, image_height = image.size
        font = ImageFont.truetype(
            os.environ.get("FONT_PATH", "/Library/Fonts/arial.ttf"),
            image_font_size)

        text_width, text_height = draw.textsize(text, font=font)
        draw.text(((image_width-text_width)/2, (image_height-text_height)/2),
                  text, fill=font_color, font=font)
        image.save(new_image_path, image_type)


# need to add test
def open_with_csv(filename, delimiter='\t'):
    data = []
    with open(filename, encoding='utf-8') as file:
        reader = csv.reader(file, delimiter=delimiter)
        for line in reader:
            data.append(line)

    return data


def write_to_csv(filename, data):
    file = csv.writer(open(filename, 'w', encoding='utf-8'), dialect='excel')
    file.writerows(data)


# need to add test
def load_data_numpy(filename, fieldnames, datatypes, delimiter='\t'):
    data = numpy.genfromtxt(filename,
                            delimiter=delimiter,
                            skip_header=1,
                            invalid_raise=False,
                            names=fieldnames,
                            dtype=datatypes)
    return data


def write_to_excel(filename, data):
    workbook = Workbook()
    worksheet = workbook.active

    row_index = 1
    for row in data:
        column_index = 1
        for column in row:
            column_letter = get_column_letter(column_index)
            worksheet.cell('{}{}'.format(column_letter, row_index)).value = \
                column
            column_index += 1
        row_index += 1

    workbook.save(filename)


# need to add test
def occurrences_in_nested_list(nested_list, search_term):
    count = 0

    for row in nested_list:
        count += row.count(search_term)

    return count


# need to add test
def import_google_spreadsheet():
    pass
