import logging
import datetime
import os
import csv

import smtplib
import numpy
from PIL import ImageFont
from PIL import Image
from PIL import ImageDraw
from openpyxl import Workbook
from openpyxl.cell import get_column_letter




# need to add test
def open_with_csv(filename, delimiter='\t'):
    data = []
    with open(filename, encoding='utf-8') as file:
        reader = csv.reader(file, delimiter=delimiter)
        for line in reader:
            data.append(line)

    return data


def write_to_csv(filename, data):
    file = csv.writer(open(filename, 'w', encoding='utf-8'), dialect='excel')
    file.writerows(data)


# need to add test
def load_data_numpy(filename, fieldnames, datatypes, delimiter='\t'):
    data = numpy.genfromtxt(filename,
                            delimiter=delimiter,
                            skip_header=1,
                            invalid_raise=False,
                            names=fieldnames,
                            dtype=datatypes)
    return data


def write_to_excel(filename, data):
    workbook = Workbook()
    worksheet = workbook.active

    row_index = 1
    for row in data:
        column_index = 1
        for column in row:
            column_letter = get_column_letter(column_index)
            worksheet.cell('{}{}'.format(column_letter, row_index)).value = \
                column
            column_index += 1
        row_index += 1

    workbook.save(filename)


# need to add test
def occurrences_in_nested_list(nested_list, search_term):
    count = 0

    for row in nested_list:
        count += row.count(search_term)

    return count


# need to add test
def import_google_spreadsheet():
    pass
